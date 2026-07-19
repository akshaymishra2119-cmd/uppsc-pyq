"""
MPPSC Prelims GS Paper 1 — PDF Downloader + Question Extractor
================================================================
Sources: prepp.in CDN & collegedunia CDN (direct PDF links, no auth needed)
Output : MPPSC_PYQ.xlsx  (same column layout as BPSC_PYQ.xlsx)

Usage:
  pip install pdfplumber openpyxl requests
  python extract_mppsc.py

For scanned PDFs (most years), also install:
  pip install pytesseract pdf2image pillow
  Then install Tesseract OCR binary for Windows:
    https://github.com/UB-Mannheim/tesseract/wiki
    → Download tesseract-ocr-w64-setup-*.exe
    → During install, expand "Additional language data" and tick "Hindi"
    → Default install path: C:\\Program Files\\Tesseract-OCR\\tesseract.exe
"""

import os, re, time, requests, openpyxl, io
import pdfplumber

OUTPUT  = r'D:\uppsc_pyq\MPPSC_PYQ_final.xlsx'
PDF_DIR = r'D:\uppsc_pyq\mppsc_pdfs'   # downloaded PDFs cached here

# ── Tesseract path (Windows default) ─────────────────────────────────────────
TESSERACT_CMD = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

os.makedirs(PDF_DIR, exist_ok=True)

# ── PDF sources ──────────────────────────────────────────────────────────────
# Only Prelims Paper 1 (General Studies) — the 100-question MCQ paper
PAPERS = [
    # (year, label, url)
    # 2026: QP is behind a JS page; answer key PDF has questions printed too
    (2026, '2026', 'https://cdn-images.prepp.in/public/image/MPPSC_Prelims_2026_GS_Answer_Key_PDF_Apr_26_2026__bb50bc2bea961ac4f88d5fe2f4028177.pdf'),
    (2025, '2025', 'https://cdn-images.prepp.in/public/image/MPPSC_2025_GS_Paper_I_Question_Paper_and_Answer_Key_PDF_f12cf7da34097a90c3d28ece94712920.pdf'),
    (2024, '2024_SetA', 'https://cdn-images.prepp.in/public/image/MPPSC_2024_Prelims_Paper_1_Set_A_59113c9fc77dfe7cea0e3b2a6258e6b3.pdf'),
    (2024, '2024_SetB', 'http://cdn-images.prepp.in/public/image/Mppsc_PAper_1_Set_B_compressed_56d6a9a86223bcfc6c45ccdefe771aaa.pdf'),
    (2023, '2023_Dec_EN', 'https://cdn-images.prepp.in/public/image/MPPSC_Prelims_Paper_1_December_17_2023_Question_Paper_Eng_Medium__e316e70ba58724ac452c5caf22a795d4.pdf'),
    (2021, '2021', 'https://cdn-images.prepp.in/public/image/Paper_I_SSE_2021_6974107edaf0263371a98f6135b7fa6a.pdf'),
    (2020, '2020_SetA', 'https://images.collegedunia.com/public/image/8bed994a82e219fab0dc455c09ea6192.pdf'),
    (2020, '2020_SetB', 'https://images.collegedunia.com/public/image/973528a7ce60589dc4bc7d771f0d8cab.pdf'),
    (2019, '2019', 'https://images.collegedunia.com/public/image/9aa045b5e0a8c70e093ee69b92fa3922.pdf'),
    (2018, '2018', 'https://images.collegedunia.com/public/image/968576f4db2e2b089184414db6736953.pdf'),
    (2017, '2017', 'https://images.collegedunia.com/public/image/7880cb8e172bce13b46008b26376a36d.pdf'),
    (2016, '2016', 'https://images.collegedunia.com/public/image/cdafd9ad4b306df6ddf39bac84d9cb18.pdf'),
    (2015, '2015', 'https://images.collegedunia.com/public/image/d7ce720625ccf23cd1a9f61438aa10bd.pdf'),
    (2014, '2014', 'https://cdn-images.prepp.in/public/image/MPPSC_General_Studies_2014_English__4ddfa15084a3fa3bcd8e8e455d532b47.pdf'),
    (2013, '2013', 'https://cdn-images.prepp.in/public/image/MPPSC_General_Studies_2013_English_c072b04150a02244c64beededad7e384.pdf'),
]

HEADERS = [
    'Q_ID', 'Year', 'Exam', 'Subject', 'Sub_Topic',
    'Question', 'Option_A', 'Option_B', 'Option_C', 'Option_D',
    'Correct_Answer', 'Correct_Option_Text', 'Explanation', 'Difficulty', 'Source'
]


# ── Download ─────────────────────────────────────────────────────────────────
def download_pdf(label, url):
    path = os.path.join(PDF_DIR, f'MPPSC_{label}.pdf')
    if os.path.exists(path):
        print(f'  [cached] {label}')
        return path
    print(f'  [download] {label}...', end=' ', flush=True)
    try:
        r = requests.get(url, timeout=60, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if r.status_code == 200 and b'%PDF' in r.content[:10]:
            with open(path, 'wb') as f:
                f.write(r.content)
            print(f'OK ({len(r.content)//1024} KB)')
            return path
        else:
            print(f'FAILED (HTTP {r.status_code})')
            return None
    except Exception as e:
        print(f'ERROR: {e}')
        return None


# ── Text extraction ───────────────────────────────────────────────────────────
def extract_text_pdfplumber(path):
    """Extract all text from a text-based PDF."""
    texts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    return '\n'.join(texts)


def is_scanned(text):
    """Heuristic: if we got very little text, PDF is probably image-based."""
    return len(text.strip()) < 200


def ocr_page_img(img, lang):
    """
    MPPSC papers are bilingual 2-column: Hindi left | English right.
    Strategy:
      1. Try full page first — if it yields clean English (≥70% ASCII words), use it.
      2. Otherwise crop to right half and OCR again.
      3. As last resort crop left half (some older papers flip the columns).
    """
    import pytesseract

    def ascii_ratio(text):
        words = text.split()
        if not words:
            return 0
        clean = sum(1 for w in words if all(ord(c) < 128 for c in w))
        return clean / len(words)

    def ocr(image):
        return pytesseract.image_to_string(image, lang=lang,
               config='--psm 4')   # psm 4 = single column of text

    # Full page
    full = ocr(img)
    if ascii_ratio(full) >= 0.70:
        return full

    # Right half
    w, h = img.size
    right_half = img.crop((w // 2, 0, w, h))
    right_text = ocr(right_half)
    if ascii_ratio(right_text) >= 0.65:
        return right_text

    # Left half fallback
    left_half = img.crop((0, 0, w // 2, h))
    left_text = ocr(left_half)
    if ascii_ratio(left_text) >= 0.65:
        return left_text

    # Last resort: return whichever half had more ASCII
    return right_text if ascii_ratio(right_text) >= ascii_ratio(left_text) else left_text


def extract_text_ocr(path):
    """OCR using pypdfium2 (already installed) + pytesseract. No poppler needed."""
    try:
        import pypdfium2 as pdfium
        import pytesseract
        from PIL import Image

        # Point pytesseract at the Windows Tesseract binary
        if os.path.exists(TESSERACT_CMD):
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

        ocr_lang = 'eng'  # English only — Hindi column will be cropped out
        print(f'    [OCR col-split]', end=' ', flush=True)

        doc = pdfium.PdfDocument(path)
        texts = []
        for i in range(len(doc)):
            page = doc[i]
            bitmap = page.render(scale=3.0)   # high DPI for better OCR
            img = bitmap.to_pil()
            t = ocr_page_img(img, ocr_lang)
            texts.append(t)
            print(f'p{i+1}', end=' ', flush=True)
        print()
        return '\n'.join(texts)

    except ImportError as e:
        print(f'\n  [!] Missing library: {e}')
        print('      Run: pip install pytesseract pillow')
        return ''
    except Exception as e:
        print(f'\n  [!] OCR error: {e}')
        return ''


# ── Clean garbled bilingual OCR output ───────────────────────────────────────
def clean_ocr_text(text):
    """
    Bilingual MPPSC papers have Hindi (left col) + English (right col).
    OCR without Hindi tessdata turns Hindi into garbage like 'aqede' / 'aRa'.
    Strategy: keep only lines where ≥60% of chars are printable ASCII.
    This filters out the Hindi column noise and keeps clean English lines.
    """
    clean_lines = []
    for line in text.split('\n'):
        if not line.strip():
            clean_lines.append('')
            continue
        total = len(line)
        ascii_count = sum(1 for c in line if ord(c) < 128 and c.isprintable())
        if total == 0 or (ascii_count / total) >= 0.60:
            clean_lines.append(line)
    return '\n'.join(clean_lines)


# ── Question parsing ──────────────────────────────────────────────────────────
# MPPSC papers typically look like:
#   1. Which of the following...?
#   (a) Option A   (b) Option B
#   (c) Option C   (d) Option D
#
# Or bilingual:
#   1. Question text in Hindi
#      Question text in English  (or vice versa)
#   (A) ...   (B) ...   (C) ...   (D) ...

Q_NUM_RE   = re.compile(r'^\s*(\d{1,3})[.\)]\s+(.+)', re.MULTILINE)
OPT_BLOCK  = re.compile(
    r'\(([aAbBcCdD1234])\)\s*(.+?)(?=\s*\([aAbBcCdD1234]\)|\Z)',
    re.DOTALL
)
# Answer key line: "1.(b) 2.(a) 3.(d)..."
ANS_LINE   = re.compile(r'(\d+)\.\s*\(?([aAbBcCdD])\)?')


def clean(s):
    if not s:
        return ''
    return re.sub(r'\s+', ' ', str(s)).strip()


def normalise_opt_key(k):
    """Map a/b/c/d or 1/2/3/4 → A/B/C/D"""
    m = {'a':'A','b':'B','c':'C','d':'D',
         '1':'A','2':'B','3':'C','4':'D'}
    return m.get(k.lower(), k.upper())


def parse_questions(text, year, label):
    """
    Parse MCQs from extracted text.
    Returns list of dicts matching HEADERS.
    """
    rows = []
    lines = text.split('\n')

    # --- Try to find answer key section ---
    answer_key = {}
    for i, line in enumerate(lines):
        matches = ANS_LINE.findall(line)
        if len(matches) >= 5:   # looks like an answer key line
            for qno, opt in matches:
                answer_key[int(qno)] = normalise_opt_key(opt)

    # --- Locate question blocks ---
    # Join lines smartly: a new question starts with "N." at line start
    full_text = '\n'.join(lines)

    # Split on question number pattern
    parts = re.split(r'\n(?=\s*\d{1,3}[.\)]\s)', full_text)

    q_counter = 0
    for part in parts:
        part = part.strip()
        if not part:
            continue

        # First line: question number + start of question text
        first_line_m = re.match(r'^(\d{1,3})[.\)]\s*(.*)', part, re.DOTALL)
        if not first_line_m:
            continue

        q_num = int(first_line_m.group(1))
        if q_num < 1 or q_num > 200:
            continue

        rest = first_line_m.group(2)

        # Extract options (a)/(b)/(c)/(d)
        opt_matches = list(OPT_BLOCK.finditer(rest))
        if len(opt_matches) < 2:
            # Try options on the full part
            opt_matches = list(OPT_BLOCK.finditer(part))

        opts = {}
        for m in opt_matches:
            key = normalise_opt_key(m.group(1))
            opts[key] = clean(m.group(2))

        # Question text = everything before first option marker
        if opt_matches:
            q_text = clean(rest[:opt_matches[0].start()])
        else:
            # No options found — might be scanned noise, skip
            continue

        # Remove trailing option junk from question text
        q_text = re.sub(r'\([aAbBcCdD1234]\)\s*$', '', q_text).strip()

        if len(q_text) < 5:
            continue

        correct_key = answer_key.get(q_num, '')
        correct_text = opts.get(correct_key, '')

        q_counter += 1
        qid = f'MPPSC_{year}_{label}_{q_num:03d}'

        rows.append({
            'Q_ID': qid,
            'Year': year,
            'Exam': 'MPPSC',
            'Subject': 'General Studies',
            'Sub_Topic': '',
            'Question': q_text,
            'Option_A': opts.get('A', ''),
            'Option_B': opts.get('B', ''),
            'Option_C': opts.get('C', ''),
            'Option_D': opts.get('D', ''),
            'Correct_Answer': correct_key,
            'Correct_Option_Text': correct_text,
            'Explanation': '',
            'Difficulty': 'Medium',
            'Source': f'MPPSC Prelims {year} Paper 1 GS ({label})'
        })

    return rows


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    all_rows = []

    print('=== MPPSC Question Extractor ===\n')

    for year, label, url in PAPERS:
        print(f'\n[{label}] {url[:60]}...')

        pdf_path = download_pdf(label, url)
        if not pdf_path:
            print(f'  Skipping {label} — download failed.')
            continue

        print(f'  Extracting text...', end=' ')
        text = extract_text_pdfplumber(pdf_path)

        used_ocr = False
        if is_scanned(text):
            print('scanned PDF detected, trying OCR...')
            text = extract_text_ocr(pdf_path)
            used_ocr = True
        else:
            words = len(text.split())
            print(f'{words} words extracted')


        if not text.strip():
            print(f'  No text extracted for {label}, skipping.')
            continue

        rows = parse_questions(text, year, label)
        print(f'  Parsed {len(rows)} questions')
        all_rows.extend(rows)

        time.sleep(0.5)  # be polite

    # ── Write Excel ──────────────────────────────────────────────────────────
    print(f'\n\nWriting {len(all_rows)} total questions to {OUTPUT}...')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MPPSC_PYQ'

    ws.append(HEADERS)
    for row in all_rows:
        ws.append([row.get(h, '') for h in HEADERS])

    # Basic formatting
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Safe save — handle file locked in Excel
    try:
        wb.save(OUTPUT)
        print(f'Done! -> {OUTPUT}')
    except PermissionError:
        alt = OUTPUT.replace('.xlsx', '_new.xlsx')
        wb.save(alt)
        print(f'  [!] {OUTPUT} is open in Excel — saved to {alt} instead.')
        print(f'      Close Excel and rename _new.xlsx → MPPSC_PYQ.xlsx')
    print(f'\nSummary by year:')
    from collections import Counter
    c = Counter(r['Year'] for r in all_rows)
    for yr in sorted(c):
        print(f'  {yr}: {c[yr]} questions')


if __name__ == '__main__':
    main()

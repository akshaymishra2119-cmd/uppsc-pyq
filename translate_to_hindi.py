"""
Translate UPPSC_PYQ.xlsx English questions to Hindi.
- Uses deep-translator (Google Translate, FREE, no API key)
- Keeps all non-text columns as-is
- Outputs: UPPSC_PYQ_Hindi.xlsx

Usage:
  pip install deep-translator openpyxl
  python translate_to_hindi.py
"""

import time
import openpyxl
from deep_translator import GoogleTranslator

INPUT  = r'D:\uppsc_pyq\UPPSC_PYQ.xlsx'
OUTPUT = r'D:\uppsc_pyq\UPPSC_PYQ_Hindi.xlsx'

# Columns to translate
TRANSLATE_COLS = ['Question', 'Option_A', 'Option_B', 'Option_C', 'Option_D',
                  'Correct_Option_Text', 'Explanation', 'Sub_Topic']

# Sub_Topic Hindi map (instant, no API call needed)
SUBTOPIC_HI = {
  'Revolt of 1857':'1857 का विद्रोह','Gandhian Movements':'गांधीवादी आंदोलन',
  'Indian National Congress':'भारतीय राष्ट्रीय कांग्रेस','Revolutionary Movements':'क्रांतिकारी आंदोलन',
  'British Administration':'ब्रिटिश प्रशासन','Social Reform Movements':'सामाजिक सुधार आंदोलन',
  'Partition & Independence':'विभाजन और स्वतंत्रता','Press & Literature':'प्रेस और साहित्य',
  'Peasant & Tribal Movements':'किसान और आदिवासी आंदोलन','Constitutional Developments':'संवैधानिक विकास',
  'Early British Conquest':'प्रारंभिक ब्रिटिश विजय','Indus Valley Civilization':'सिंधु घाटी सभ्यता',
  'Vedic Age':'वैदिक काल','Buddhism':'बौद्ध धर्म','Jainism':'जैन धर्म',
  'Mauryan Empire':'मौर्य साम्राज्य','Gupta Empire':'गुप्त साम्राज्य',
  'Post-Gupta Period':'गुप्तोत्तर काल','Medieval Kingdoms':'मध्यकालीन राज्य',
  'Constitution & Constituent Assembly':'संविधान और संविधान सभा',
  'Fundamental Rights & Duties':'मौलिक अधिकार और कर्तव्य',
  'Directive Principles':'नीति निदेशक तत्व','Parliament':'संसद',
  'President & Vice President':'राष्ट्रपति और उप राष्ट्रपति',
  'Prime Minister & Cabinet':'प्रधानमंत्री और मंत्रिमंडल',
  'Supreme Court & Judiciary':'सर्वोच्च न्यायालय और न्यायपालिका',
  'Governor & State Legislature':'राज्यपाल और राज्य विधानमंडल',
  'Elections & ECI':'चुनाव और निर्वाचन आयोग',
  'Panchayati Raj & Municipalities':'पंचायती राज और नगरपालिका',
  'Constitutional Amendments':'संवैधानिक संशोधन','Emergency Provisions':'आपातकालीन प्रावधान',
  'Schedules & Lists':'अनुसूचियाँ और सूचियाँ','Rivers & Water Bodies':'नदियाँ और जल निकाय',
  'Mountains & Passes':'पर्वत और दर्रे','Climate & Monsoon':'जलवायु और मानसून',
  'Soils & Agriculture':'मिट्टी और कृषि','National Parks & Wildlife':'राष्ट्रीय उद्यान और वन्यजीव',
  'Minerals & Energy':'खनिज और ऊर्जा','Indian States & Cities':'भारतीय राज्य और शहर',
  'World Geography':'विश्व भूगोल','Transport & Industry':'परिवहन और उद्योग',
  'Planning & Five Year Plans':'योजना और पंचवर्षीय योजनाएँ','Banking & Finance':'बैंकिंग और वित्त',
  'Budget & Fiscal Policy':'बजट और राजकोषीय नीति',
  'Agriculture & Rural Economy':'कृषि और ग्रामीण अर्थव्यवस्था',
  'Trade & Industry':'व्यापार और उद्योग','Poverty & Employment':'गरीबी और रोजगार',
  'Schemes & Programmes':'योजनाएँ और कार्यक्रम','Physics':'भौतिकी','Chemistry':'रसायन विज्ञान',
  'Biology & Health':'जीव विज्ञान और स्वास्थ्य','Technology & Computers':'प्रौद्योगिकी और कंप्यूटर',
  'Climate Change & Pollution':'जलवायु परिवर्तन और प्रदूषण',
  'Ecology & Biodiversity':'पारिस्थितिकी और जैव विविधता',
  'International Affairs':'अंतर्राष्ट्रीय मामले','Awards & Sports':'पुरस्कार और खेल',
  'Appointments & Reports':'नियुक्तियाँ और रिपोर्ट','General Studies':'सामान्य अध्ययन',
  'UP History':'उत्तर प्रदेश का इतिहास','UP Geography':'उत्तर प्रदेश का भूगोल',
  'UP Economy':'उत्तर प्रदेश की अर्थव्यवस्था','UP Polity':'उत्तर प्रदेश की राजव्यवस्था',
  'UP Culture':'उत्तर प्रदेश की संस्कृति','Ancient History':'प्राचीन इतिहास',
  'Medieval History':'मध्यकालीन इतिहास','Modern History':'आधुनिक इतिहास',
  'Environment':'पर्यावरण','Science':'विज्ञान','Economy':'अर्थव्यवस्था',
  'History':'इतिहास','Geography':'भूगोल','Polity':'राजव्यवस्था',
  'Current Affairs':'समसामयिकी','UP Special':'UP विशेष',
}

def tr(text):
    """Translate English text to Hindi, with retry."""
    if not text or not str(text).strip():
        return text
    try:
        result = GoogleTranslator(source='en', target='hi').translate(str(text).strip())
        time.sleep(0.35)
        return result
    except Exception as e:
        print(f'  [warn] {e} — retrying in 3s...')
        time.sleep(3)
        try:
            return GoogleTranslator(source='en', target='hi').translate(str(text).strip())
        except:
            return text  # fallback: keep original

def main():
    print(f'Loading {INPUT}...')
    wb_in = openpyxl.load_workbook(INPUT)
    ws_in = wb_in.active

    # Build header -> col index map
    headers = [ws_in.cell(1, c).value for c in range(1, ws_in.max_column + 1)]
    col_map = {h: i+1 for i, h in enumerate(headers) if h}

    total = ws_in.max_row - 1
    print(f'Found {total} questions. Starting translation...\n')

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Question_Bank_HI'

    # Write headers
    for c, h in enumerate(headers, 1):
        ws_out.cell(1, c, h)

    # Process each row
    for r in range(2, ws_in.max_row + 1):
        qid = ws_in.cell(r, col_map.get('Q_ID', 1)).value
        if not qid:
            continue

        print(f'[{r-1}/{total}] {qid}...', end=' ', flush=True)

        for c in range(1, ws_in.max_column + 1):
            header = headers[c-1]
            val = ws_in.cell(r, c).value

            if header == 'Sub_Topic' and val:
                # Use lookup first, fall back to translate
                ws_out.cell(r, c, SUBTOPIC_HI.get(str(val), tr(val)))

            elif header in TRANSLATE_COLS and val:
                ws_out.cell(r, c, tr(str(val)))

            else:
                # Keep as-is: Q_ID, Year, Subject, Correct_Answer, Difficulty, etc.
                ws_out.cell(r, c, val)

        print('✓')

        # Checkpoint save every 50 rows
        if (r - 1) % 50 == 0:
            wb_out.save(OUTPUT)
            print(f'  ── checkpoint saved at {r-1}/{total} ──')
            time.sleep(2)

    wb_out.save(OUTPUT)
    print(f'\nDone! -> {OUTPUT}')

if __name__ == '__main__':
    main()

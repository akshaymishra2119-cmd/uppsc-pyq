"""
UPSC CSE Prelims GS Paper 1 — Extractor v3
==========================================
- Text PDFs (2023): pdfplumber right-column crop
- Scanned PDFs (2018-2022): pypdfium2 + pytesseract OCR on right half
- Answer keys: hardcoded from official UPSC published keys (2018-2022)
                + URL fetch attempt for 2023
"""

import os, re, time, requests, openpyxl
import pdfplumber
import pypdfium2 as pdfium
from PIL import Image
import numpy as np
import cv2
import pytesseract

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

OUTPUT  = r'D:\uppsc_pyq\UPSC_PYQ.xlsx'
PDF_DIR = r'D:\uppsc_pyq\upsc_pdfs'
os.makedirs(PDF_DIR, exist_ok=True)

HEADERS = ['Q_ID','Year','Exam','Subject','Sub_Topic',
           'Question','Option_A','Option_B','Option_C','Option_D',
           'Correct_Answer','Correct_Option_Text','Explanation','Difficulty','Source']

# ── Official UPSC Set-A answer keys (read from official UPSC-published PDFs) ──
HARDCODED_AK = {
    2022: {
         1:'B',  2:'C',  3:'B',  4:'C',  5:'A',  6:'D',  7:'A',  8:'D',  9:'A', 10:'C',
        11:'B', 12:'B', 13:'B', 14:'B', 15:'B', 16:'B', 17:'D', 18:'D', 19:'B', 20:'A',
        21:'B', 22:'D', 23:'B', 24:'C', 25:'B', 26:'A', 27:'C', 28:'B', 29:'B', 30:'B',
        31:'D', 32:'D', 33:'D', 34:'C', 35:'B', 36:'D', 37:'D', 38:'C', 39:'B', 40:'C',
        41:'A', 42:'B', 43:'D', 44:'B', 45:'C', 46:'A', 47:'A', 48:'A', 49:'A', 50:'C',
        51:'C', 52:'B', 53:'D', 54:'B', 55:'B', 56:'C', 57:'B', 58:'D', 59:'B', 60:'B',
        61:'X', 62:'D', 63:'B', 64:'B', 65:'C', 66:'D', 67:'C', 68:'D', 69:'A', 70:'C',
        71:'C', 72:'A', 73:'A', 74:'D', 75:'C', 76:'A', 77:'A', 78:'D', 79:'D', 80:'A',
        81:'D', 82:'C', 83:'C', 84:'B', 85:'D', 86:'B', 87:'C', 88:'B', 89:'B', 90:'A',
        91:'B', 92:'B', 93:'B', 94:'A', 95:'A', 96:'A', 97:'D', 98:'D', 99:'D',100:'B',
    },
    2021: {
         1:'C',  2:'B',  3:'B',  4:'A',  5:'B',  6:'D',  7:'A',  8:'A',  9:'D', 10:'D',
        11:'C', 12:'A', 13:'B', 14:'C', 15:'B', 16:'A', 17:'B', 18:'D', 19:'A', 20:'C',
        21:'C', 22:'B', 23:'D', 24:'A', 25:'B', 26:'C', 27:'C', 28:'C', 29:'A', 30:'D',
        31:'C', 32:'A', 33:'A', 34:'B', 35:'D', 36:'C', 37:'D', 38:'A', 39:'C', 40:'B',
        41:'B', 42:'B', 43:'A', 44:'C', 45:'A', 46:'C', 47:'D', 48:'B', 49:'A', 50:'B',
        51:'B', 52:'B', 53:'D', 54:'D', 55:'B', 56:'B', 57:'A', 58:'C', 59:'D', 60:'D',
        61:'C', 62:'B', 63:'B', 64:'B', 65:'C', 66:'C', 67:'B', 68:'C', 69:'A', 70:'B',
        71:'A', 72:'C', 73:'D', 74:'B', 75:'D', 76:'D', 77:'C', 78:'B', 79:'C', 80:'X',
        81:'D', 82:'B', 83:'B', 84:'D', 85:'A', 86:'A', 87:'A', 88:'A', 89:'A', 90:'D',
        91:'B', 92:'B', 93:'B', 94:'D', 95:'D', 96:'D', 97:'D', 98:'C', 99:'B',100:'D',
    },
    2020: {
         1:'B',  2:'B',  3:'D',  4:'D',  5:'B',  6:'D',  7:'D',  8:'D',  9:'A', 10:'C',
        11:'B', 12:'A', 13:'D', 14:'A', 15:'D', 16:'D', 17:'D', 18:'D', 19:'D', 20:'C',
        21:'B', 22:'B', 23:'A', 24:'C', 25:'C', 26:'A', 27:'X', 28:'A', 29:'A', 30:'A',
        31:'C', 32:'B', 33:'B', 34:'D', 35:'D', 36:'C', 37:'D', 38:'B', 39:'C', 40:'D',
        41:'C', 42:'D', 43:'D', 44:'D', 45:'B', 46:'A', 47:'C', 48:'A', 49:'A', 50:'D',
        51:'B', 52:'X', 53:'A', 54:'D', 55:'B', 56:'C', 57:'B', 58:'B', 59:'B', 60:'B',
        61:'C', 62:'A', 63:'D', 64:'B', 65:'A', 66:'B', 67:'A', 68:'C', 69:'D', 70:'C',
        71:'B', 72:'A', 73:'C', 74:'A', 75:'A', 76:'D', 77:'A', 78:'A', 79:'D', 80:'D',
        81:'A', 82:'A', 83:'D', 84:'A', 85:'A', 86:'A', 87:'A', 88:'D', 89:'C', 90:'A',
        91:'C', 92:'D', 93:'B', 94:'B', 95:'C', 96:'D', 97:'A', 98:'B', 99:'C',100:'C',
    },
    2019: {
         1:'D',  2:'B',  3:'C',  4:'A',  5:'C',  6:'D',  7:'C',  8:'A',  9:'D', 10:'A',
        11:'D', 12:'A', 13:'D', 14:'B', 15:'D', 16:'A', 17:'C', 18:'D', 19:'C', 20:'A',
        21:'A', 22:'A', 23:'A', 24:'B', 25:'D', 26:'D', 27:'A', 28:'D', 29:'C', 30:'A',
        31:'D', 32:'D', 33:'C', 34:'D', 35:'D', 36:'B', 37:'B', 38:'A', 39:'A', 40:'C',
        41:'D', 42:'D', 43:'B', 44:'B', 45:'B', 46:'C', 47:'A', 48:'A', 49:'A', 50:'B',
        51:'C', 52:'B', 53:'C', 54:'C', 55:'B', 56:'B', 57:'C', 58:'B', 59:'C', 60:'C',
        61:'B', 62:'A', 63:'D', 64:'B', 65:'B', 66:'C', 67:'D', 68:'A', 69:'A', 70:'A',
        71:'A', 72:'D', 73:'A', 74:'B', 75:'D', 76:'C', 77:'A', 78:'C', 79:'C', 80:'D',
        81:'D', 82:'A', 83:'A', 84:'D', 85:'D', 86:'D', 87:'A', 88:'C', 89:'B', 90:'B',
        91:'B', 92:'B', 93:'A', 94:'C', 95:'D', 96:'A', 97:'B', 98:'B', 99:'A',100:'B',
    },
    2018: {
         1:'B',  2:'D',  3:'C',  4:'B',  5:'A',  6:'C',  7:'A',  8:'D',  9:'C', 10:'A',
        11:'C', 12:'D', 13:'A', 14:'D', 15:'C', 16:'A', 17:'C', 18:'D', 19:'A', 20:'B',
        21:'B', 22:'B', 23:'A', 24:'B', 25:'B', 26:'D', 27:'D', 28:'A', 29:'A', 30:'A',
        31:'C', 32:'A', 33:'A', 34:'C', 35:'B', 36:'B', 37:'B', 38:'B', 39:'A', 40:'B',
        41:'C', 42:'D', 43:'C', 44:'B', 45:'C', 46:'B', 47:'C', 48:'C', 49:'C', 50:'D',
        51:'C', 52:'C', 53:'B', 54:'A', 55:'D', 56:'C', 57:'C', 58:'C', 59:'C', 60:'D',
        61:'A', 62:'D', 63:'B', 64:'B', 65:'A', 66:'B', 67:'D', 68:'D', 69:'C', 70:'D',
        71:'B', 72:'D', 73:'B', 74:'C', 75:'C', 76:'D', 77:'A', 78:'C', 79:'A', 80:'B',
        81:'B', 82:'B', 83:'B', 84:'D', 85:'A', 86:'B', 87:'D', 88:'C', 89:'B', 90:'C',
        91:'B', 92:'C', 93:'B', 94:'A', 95:'A', 96:'B', 97:'C', 98:'A', 99:'D',100:'B',
    },
}

PAPERS = [
    (2023,
     'https://cdn-images.prepp.in/public/image/UPSC_GS_Paper_1_SET_A_Question_Paper_PDF_eed6418a335eb05f508949b024c600f0.pdf',
     # Try multiple AK URLs for 2023
     ['https://cdn-images.prepp.in/public/image/upsc-prelims-2023-answer-key-gs-paper-1-set-a.pdf',
      'https://cdn-images.prepp.in/public/image/UPSC_Prelims_2023_GS1_Answer_Key_Set_A.pdf']),
    (2022,
     'https://cdn-images.prepp.in/public/image/1ef88a261b212cc106dc0b6ae01cee9d.pdf', None),
    (2021,
     'https://cdn-images.prepp.in/public/image/979a60c746ad97ef11224f59cfbd3608.pdf', None),
    (2020,
     'https://cdn-images.prepp.in/public/image/a5770d494930b50bb2fe452397a4eff6.pdf', None),
    (2019,
     'https://cdn-images.prepp.in/public/image/e1820086e52cd1cfdce5b729fd17050a.pdf', None),
    (2018,
     'https://cdn-images.prepp.in/public/image/UPSC_Prelims_2018_June_03_GS_I_Paper_I_Question_Paper_and_Answer_Key_PDF_b8620a03659801b5cd11b01382ec6b06.pdf',
     None),
]

SESS = requests.Session()
SESS.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
SESS.headers['Referer']    = 'https://prepp.in/'


def download(fname, url):
    path = os.path.join(PDF_DIR, fname)
    if os.path.exists(path):
        print(f'  [cached] {fname}')
        return path
    print(f'  [dl] {fname}...', end=' ', flush=True)
    try:
        r = SESS.get(url, timeout=60)
        if r.status_code == 200 and r.content[:4] == b'%PDF':
            open(path, 'wb').write(r.content)
            print(f'OK {len(r.content)//1024}KB')
            return path
        print(f'FAIL {r.status_code}')
        return None
    except Exception as e:
        print(f'ERR {e}')
        return None


def extract_english_text(path):
    """Text-based PDF: crop right 50% with pdfplumber."""
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            w, h = page.width, page.height
            col = page.crop((w * 0.50, 0, w, h))
            t = col.extract_text()
            if t:
                pages.append(t)
    return '\n'.join(pages)


def extract_english_ocr(path):
    """Scanned PDF: render pages, crop right half, OCR."""
    doc = pdfium.PdfDocument(path)
    pages = []
    for i in range(len(doc)):
        page = doc[i]
        bm = page.render(scale=3)
        img = bm.to_pil().convert('L')
        w, h = img.size
        # Crop right ~52% (English column, with small overlap buffer)
        right = img.crop((int(w * 0.48), 0, w, h))
        # Sharpen via threshold
        arr = np.array(right)
        _, arr = cv2.threshold(arr, 150, 255, cv2.THRESH_BINARY)
        right = Image.fromarray(arr)
        text = pytesseract.image_to_string(right, lang='eng', config='--psm 6')
        if text.strip():
            pages.append(text)
    doc.close()
    return '\n'.join(pages)


def get_text(path, year):
    """Auto-detect text vs scanned and extract English column."""
    # Quick probe: can pdfplumber get meaningful text?
    text = extract_english_text(path)
    words = len(text.split())
    if words >= 200:
        print(f'text-based ({words}w)')
        return text
    print(f'scanned ({words}w) → OCR...', end=' ', flush=True)
    text = extract_english_ocr(path)
    print(f'{len(text.split())}w')
    return text


def parse_ak_text(text):
    """Parse answer key from text (for years we try to download AK)."""
    ak = {}
    for m in re.finditer(r'\b(\d{1,3})\s*[.):\s]\s*([abcdABCD])\b', text):
        n = int(m.group(1))
        if 1 <= n <= 100:
            ak[n] = m.group(2).upper()
    return ak


def clean(s):
    return re.sub(r'\s+', ' ', s or '').strip()


OPT = re.compile(r'\(([abcdABCD])\)\s*(.+?)(?=\s*\([abcdABCD]\)|\Z)', re.DOTALL)


def parse_questions(text, year, ak):
    rows = []
    parts = re.split(r'\n(?=\s*\d{1,3}[.)]\s)', text)
    for part in parts:
        part = part.strip()
        m = re.match(r'^(\d{1,3})[.)]\s*(.*)', part, re.DOTALL)
        if not m:
            continue
        qno = int(m.group(1))
        if not 1 <= qno <= 100:
            continue
        body = m.group(2)
        opts_found = list(OPT.finditer(body))
        if len(opts_found) < 2:
            continue
        opts = {o.group(1).upper(): clean(o.group(2)) for o in opts_found}
        q_text = clean(body[:opts_found[0].start()])
        q_text = re.sub(r'\s*\([abcdABCD]\)\s*$', '', q_text).strip()
        if len(q_text) < 10:
            continue
        # Filter Hindi-leaked text
        if sum(1 for c in q_text if ord(c) < 128) / max(len(q_text), 1) < 0.70:
            continue
        correct = ak.get(qno, '')
        rows.append({
            'Q_ID': f'UPSC_{year}_{qno:03d}', 'Year': year,
            'Exam': 'UPSC', 'Subject': 'General Studies', 'Sub_Topic': '',
            'Question': q_text,
            'Option_A': opts.get('A', ''), 'Option_B': opts.get('B', ''),
            'Option_C': opts.get('C', ''), 'Option_D': opts.get('D', ''),
            'Correct_Answer': correct,
            'Correct_Option_Text': opts.get(correct, ''),
            'Explanation': '', 'Difficulty': 'Medium',
            'Source': f'UPSC CSE Prelims {year} GS Paper 1',
        })
    return rows


def main():
    all_rows = []
    print('=== UPSC Extractor v3 ===\n')

    for entry in PAPERS:
        year = entry[0]
        qp_url = entry[1]
        ak_urls = entry[2]  # list of URLs to try, or None

        print(f'\n[{year}]')
        qp = download(f'QP_{year}.pdf', qp_url)
        if not qp:
            continue

        # Answer key: hardcoded first, then try URL fetch for 2023
        ak = HARDCODED_AK.get(year, {})
        if not ak and ak_urls:
            for ak_url in ak_urls:
                ak_path = download(f'AK_{year}.pdf', ak_url)
                if ak_path:
                    with pdfplumber.open(ak_path) as pdf:
                        ak_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
                    ak = parse_ak_text(ak_text)
                    if ak:
                        print(f'  AK from URL: {len(ak)} answers')
                        break
            if not ak:
                print(f'  AK: not found (answers will be blank)')
        else:
            print(f'  AK: {len(ak)} answers (hardcoded)')

        print(f'  Extracting...', end=' ')
        text = get_text(qp, year)
        rows = parse_questions(text, year, ak)
        matched = sum(1 for r in rows if r['Correct_Answer'])
        print(f'  {len(rows)}q | {matched} with answers')
        all_rows.extend(rows)
        time.sleep(0.2)

    print(f'\nTotal: {len(all_rows)} questions')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'UPSC_PYQ'
    ws.append(HEADERS)
    for r in all_rows:
        ws.append([r.get(h, '') for h in HEADERS])
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = PatternFill(fill_type='solid', fgColor='203864')
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 60)
    try:
        wb.save(OUTPUT)
        print(f'Saved -> {OUTPUT}')
    except PermissionError:
        alt = OUTPUT.replace('.xlsx', '_out.xlsx')
        wb.save(alt)
        print(f'Locked — saved to {alt}')

    from collections import Counter
    for yr, cnt in sorted(Counter(r['Year'] for r in all_rows).items()):
        matched = sum(1 for r in all_rows if r['Year'] == yr and r['Correct_Answer'])
        print(f'  {yr}: {cnt}q | {matched} answered')


if __name__ == '__main__':
    main()
